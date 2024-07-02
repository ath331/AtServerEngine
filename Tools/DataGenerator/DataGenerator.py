from openpyxl import load_workbook


import argparse
import os
import glob
import jinja2


class Member:
    def __init__(self, Name, name, valueType, default):
        self.Name = Name
        self.name = name
        self.valueType = valueType
        self.default = default


def main():
    arg_parser = argparse.ArgumentParser(description = 'DataGenerator')
    arg_parser.add_argument('--path', type=str, default='../../Data', help='data dir path')
    args = arg_parser.parse_args()

    for root, dirs, files in os.walk(args.path):
        for file in files:
            if file.endswith(".xlsx"):
                file_path = os.path.join(root, file)

                memberList = []

                # xlsx
                workbook = load_workbook(file_path)

                for sheetname in workbook.sheetnames:

                    # sheet
                    sheet = workbook[sheetname]

                    # col
                    for col in sheet.iter_cols(values_only=True):
                        if all(value is None for value in col):
                            break

                        memberNAME = "NAME"
                        memberName = "name"
                        memberType = "type"
                        memberDefault = "default"

                        for index, cell in enumerate(col):

                            if index == 0:
                                memberNAME = cell
                                memberName = memberNAME[0].lower() + memberNAME[1:]
                                continue

                            if index == 3:
                                memberType = cell
                                if memberType.startswith("E"):
                                    memberType = "Protocol::" + memberType

                                memberDefault = memberType + "( 0 )"
                                if memberType == "AtString":
                                    memberDefault = "\"\""
                                if memberType.startswith("E"):
                                    memberDefault = memberType + "::Max"

                                continue


                        member = Member(memberNAME, memberName, memberType, memberDefault)

                        memberList.append(member)


                workbook.close()

            file_loader = jinja2.FileSystemLoader('Templates')
            env = jinja2.Environment(loader=file_loader)

            template = env.get_template('InfoTemplate.h')
            output = template.render(ClassName=sheetname, memberList=memberList)

            f = open(sheetname+'InfoTemplate.h', 'w+')
            f.write(output)
            f.close()

            template = env.get_template('InfoTemplate.cpp')
            output = template.render(ClassName=sheetname, memberList=memberList)

            f = open(sheetname+'InfoTemplate.cpp', 'w+')
            f.write(output)
            f.close()

            print(output)



                
    return



if __name__ == '__main__':
	main()